import csv
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# --- Load CSV ---
rows = []
with open("japan_characters_100.csv", encoding="utf-8") as f:
    reader = csv.reader(f)
    for row in reader:
        rows.append(row)

headers = rows[0]
data = rows[1:]

wb = Workbook()
ws = wb.active
ws.title = "日本企業キャラクター100事例"

# --- Styles ---
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
HEADER_FONT   = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=10)
ROW_FILL_ODD  = PatternFill("solid", fgColor="EBF3FB")
ROW_FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")
BODY_FONT     = Font(name="Meiryo UI", size=9)
WRAP_ALIGN    = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN  = Alignment(wrap_text=False, vertical="center", horizontal="center")

thin = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# --- Column widths (chars) ---
# 14 columns:
# 1:No. 2:キャラクター名 3:企業/ブランド名 4:業種/業界 5:誕生年
# 6:キャラクタータイプ 7:モチーフ 8:デザイン・ビジュアル特徴
# 9:キャラクター設定・バックストーリー 10:制作背景・誕生経緯
# 11:主な活用媒体 12:ターゲット層 13:IP展開・グッズ化 14:現在の状況
COL_WIDTHS = {
    1:  5,   # No.
    2:  20,  # キャラクター名
    3:  22,  # 企業/ブランド名
    4:  16,  # 業種/業界
    5:  8,   # 誕生年
    6:  18,  # キャラクタータイプ
    7:  16,  # モチーフ
    8:  36,  # デザイン・ビジュアル特徴
    9:  46,  # キャラクター設定・バックストーリー
    10: 46,  # 制作背景・誕生経緯
    11: 24,  # 主な活用媒体
    12: 18,  # ターゲット層
    13: 28,  # IP展開・グッズ化
    14: 12,  # 現在の状況
}

# --- Write header row ---
ws.row_dimensions[1].height = 30
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.border    = BORDER
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_idx, 15)

# --- Write data rows ---
for row_idx, row in enumerate(data, start=2):
    fill = ROW_FILL_ODD if row_idx % 2 == 0 else ROW_FILL_EVEN
    ws.row_dimensions[row_idx].height = 72

    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill   = fill
        cell.font   = BODY_FONT
        cell.border = BORDER

        if col_idx == 1:
            cell.alignment = CENTER_ALIGN
        elif col_idx in (5, 14):
            cell.alignment = Alignment(wrap_text=False, vertical="top", horizontal="center")
        else:
            cell.alignment = WRAP_ALIGN

# --- Freeze header row ---
ws.freeze_panes = "A2"

# --- Auto-filter ---
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# --- Sheet tab color ---
ws.sheet_properties.tabColor = "1F3864"

wb.save("japan_characters_100.xlsx")
print("Saved: japan_characters_100.xlsx")
